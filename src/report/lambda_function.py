import boto3
import io
import json
import os
import time
import email.mime.multipart as mp
import email.mime.text      as mt
import email.mime.application as ma
from collections import defaultdict
from datetime    import datetime, timedelta
from openpyxl                import Workbook
from openpyxl.styles         import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils          import get_column_letter
from openpyxl.chart          import BarChart, Reference

# ── 🔹 Clientes AWS ───────────────────────────────────────────────
athena     = boto3.client('athena')
ses_client = boto3.client('ses', region_name='us-east-1')
try:
    orgs = boto3.client('organizations')
except Exception:
    orgs = None

# ── 🔐 Config desde variables de entorno ──────────────────────────
DATABASE   = os.environ["ATHENA_DATABASE"]
TABLE      = os.environ["ATHENA_TABLE"]
OUTPUT     = os.environ["ATHENA_OUTPUT"]
FROM_EMAIL = os.environ["SES_FROM_EMAIL"]

# ─────────────────────────────────────────────────────────────────
# 🎨 ESTILOS EXCEL
# ─────────────────────────────────────────────────────────────────
AWS_DARK   = "232F3E"
AWS_ORANGE = "FF9900"
HEADER2    = "2C3E50"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F4F6"
LIGHT_BLUE = "EBF5FB"
LIGHT_RED  = "FDEDEC"
LIGHT_GRN  = "EAFAF1"
LIGHT_ORG  = "FEF9E7"
BORDER_C   = "D5D8DC"
GREEN_C    = "1E8449"
RED_C      = "C0392B"
AMBER_C    = "D68910"

def bdr():
    s = Side(style='thin', color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(c):   return PatternFill("solid", fgColor=c)
def hf(sz=10, color=WHITE): return Font(name='Arial', bold=True, size=sz, color=color)
def nf(sz=9, bold=False, color=AWS_DARK): return Font(name='Arial', size=sz, bold=bold, color=color)
def ctr(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def rgt(): return Alignment(horizontal='right',  vertical='center')
def lft(): return Alignment(horizontal='left',   vertical='center')

def set_header_row(ws, row, vals, bg, fsize=9, fcolor=WHITE):
    ws.row_dimensions[row].height = 22
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v)
        c.font = hf(fsize, fcolor)
        c.fill = fill(bg)
        c.alignment = ctr()
        c.border = bdr()

def set_data_row(ws, row, vals, bg, fmt_map=None):
    ws.row_dimensions[row].height = 18
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v)
        c.fill = fill(bg)
        c.border = bdr()
        c.font   = nf()
        if fmt_map and i in fmt_map:
            c.number_format = fmt_map[i]
            c.alignment = rgt()
        elif i == 1:
            c.alignment = ctr()
        else:
            c.alignment = lft()

# ─────────────────────────────────────────────────────────────────
# 🔍 ATHENA — queries sobre aws_costs.data
# ─────────────────────────────────────────────────────────────────
def run_query(sql):
    resp = athena.start_query_execution(
        QueryString=sql,
        QueryExecutionContext={'Database': DATABASE},
        ResultConfiguration={'OutputLocation': OUTPUT},
        WorkGroup='primary'
    )
    qid = resp['QueryExecutionId']

    # ⏳ Esperar resultado
    while True:
        r     = athena.get_query_execution(QueryExecutionId=qid)
        state = r['QueryExecution']['Status']['State']
        if state == 'SUCCEEDED':
            break
        if state in ['FAILED', 'CANCELLED']:
            raise Exception(f"Athena error: {r['QueryExecution']['Status'].get('StateChangeReason', 'unknown')}")
        time.sleep(2)

    # 📦 Paginar resultados
    rows, nxt = [], None
    while True:
        args = {'QueryExecutionId': qid}
        if nxt:
            args['NextToken'] = nxt
        res = athena.get_query_results(**args)
        rows.extend(res['ResultSet']['Rows'])
        nxt = res.get('NextToken')
        if not nxt:
            break
    return rows[1:]  # ⏭ omitir fila de headers

def get_account_names():
    # 🏢 Obtiene nombres reales de cuentas via AWS Organizations
    try:
        names = {}
        pag = orgs.get_paginator('list_accounts')
        for page in pag.paginate():
            for a in page['Accounts']:
                names[a['Id']] = a['Name']
        return names
    except Exception:
        return {}

def fetch_costs(days=30, start_date=None, end_date=None):
    today = datetime.utcnow()

    # 📅 Si vienen fechas exactas del calendario las usamos directamente
    if start_date and end_date:
        pass   # ya están definidas
    else:
        end_date   = today.strftime('%Y-%m-%d')
        start_date = (today - timedelta(days=days)).strftime('%Y-%m-%d')

    # 🔒 Nunca consultar antes del primer día con datos CUR en Athena
    CUR_START_DATE = "2026-03-01"
    if start_date < CUR_START_DATE:
        print(f"⚠️ start_date {start_date} ajustado al mínimo {CUR_START_DATE}")
        start_date = CUR_START_DATE

    where_base = f"""
        line_item_line_item_type = 'Usage'
        AND date(line_item_usage_start_date)
            BETWEEN DATE('{start_date}') AND DATE('{end_date}')
    """

    # 📊 Query 1: costos por servicio (top 20)
    rows_svc = run_query(f"""
        SELECT line_item_product_code,
               SUM(line_item_unblended_cost) AS total_cost
        FROM {TABLE}
        WHERE {where_base}
        GROUP BY line_item_product_code
        ORDER BY total_cost DESC
        LIMIT 20
    """)
    services = []
    for row in rows_svc:
        svc  = row['Data'][0].get('VarCharValue', 'N/A')
        cost = float(row['Data'][1].get('VarCharValue', '0') or 0)
        if cost >= 0.01:
            services.append({'service': svc, 'cost': round(cost, 2)})

    # 📊 Query 2: costos por cuenta + servicio
    rows_acc = run_query(f"""
        SELECT line_item_usage_account_id,
               line_item_product_code,
               SUM(line_item_unblended_cost) AS total_cost
        FROM {TABLE}
        WHERE {where_base}
        GROUP BY line_item_usage_account_id, line_item_product_code
        ORDER BY total_cost DESC
    """)
    by_account = defaultdict(list)
    for row in rows_acc:
        acct = row['Data'][0].get('VarCharValue', 'N/A')
        svc  = row['Data'][1].get('VarCharValue', 'N/A')
        cost = float(row['Data'][2].get('VarCharValue', '0') or 0)
        if cost >= 0.01:
            by_account[acct].append({'service': svc, 'cost': round(cost, 2)})

    # 📊 Query 3: tendencia mensual últimos 6 meses
    rows_trend = run_query(f"""
        SELECT DATE_FORMAT(date_trunc('month', date(line_item_usage_start_date)), '%Y-%m') AS mes,
               line_item_product_code,
               SUM(line_item_unblended_cost) AS total_cost
        FROM {TABLE}
        WHERE line_item_line_item_type = 'Usage'
          AND date(line_item_usage_start_date) >= DATE_ADD('month', -6, CURRENT_DATE)
        GROUP BY 1, 2
        ORDER BY 1, 3 DESC
    """)
    trend = defaultdict(lambda: defaultdict(float))
    for row in rows_trend:
        mes  = row['Data'][0].get('VarCharValue', '')
        svc  = row['Data'][1].get('VarCharValue', '')
        cost = float(row['Data'][2].get('VarCharValue', '0') or 0)
        if mes and cost >= 0.01:
            trend[mes][svc] += cost

    account_names = get_account_names()
    total = sum(s['cost'] for s in services)

    # ─────────────────────────────────────────────────────────────
    # 📊 Query 4 — Costos DIARIOS por cuenta para detectar alertas
    # ─────────────────────────────────────────────────────────────
    rows_daily = run_query(f"""
        SELECT
            line_item_usage_account_id,
            date(line_item_usage_start_date) AS dia,
            SUM(line_item_unblended_cost)    AS costo_dia
        FROM {TABLE}
        WHERE line_item_line_item_type = 'Usage'
          AND date(line_item_usage_start_date)
              BETWEEN DATE('{start_date}') AND DATE('{end_date}')
        GROUP BY line_item_usage_account_id, date(line_item_usage_start_date)
        ORDER BY line_item_usage_account_id, dia
    """)

    # 📊 Organizar costos diarios por cuenta
    daily_by_account = defaultdict(list)
    for row in rows_daily:
        acct = row['Data'][0].get('VarCharValue', 'N/A')
        dia  = row['Data'][1].get('VarCharValue', '')
        cost = float(row['Data'][2].get('VarCharValue', '0') or 0)
        if dia:
            daily_by_account[acct].append((dia, round(cost, 2)))

    # ─────────────────────────────────────────────────────────────
    # 🚨 Detectar los 2 casos de alertas (umbral $500)
    # ─────────────────────────────────────────────────────────────
    SPIKE_THRESHOLD = 500
    alerts_caso1    = []  # Acumulado total del período supera $500 → reportar día exacto del cruce
    alerts_caso2    = []  # Incremento entre 2 puntos del período supera $500 → reportar ambos días

    for acct_id, days_list in daily_by_account.items():
        name        = account_names.get(acct_id, acct_id)
        days_sorted = sorted(days_list, key=lambda x: x[0])

        # Calcular acumulado día a día
        acumulados = []  # [(fecha, costo_dia, acumulado)]
        acum = 0
        for dia, cost in days_sorted:
            acum += cost
            acumulados.append((dia, cost, round(acum, 2)))

        total_acct = acum  # acumulado final del período

        # ── CASO 1: acumulado total llega a +$500 en algún punto ──
        # Identifica el día EXACTO en que el acumulado cruzó $500
        if total_acct >= SPIKE_THRESHOLD:
            dia_cruce   = None
            acum_cruce  = 0
            for dia, cost, acum_d in acumulados:
                if acum_d >= SPIKE_THRESHOLD and not dia_cruce:
                    dia_cruce  = dia
                    acum_cruce = acum_d
            alerts_caso1.append({
                'account_id'   : acct_id,
                'account_name' : name,
                'total'        : round(total_acct, 2),
                'dia_cruce'    : dia_cruce,
                'acum_cruce'   : acum_cruce,
                'dias_totales' : len(days_sorted),
                'fecha_inicio' : days_sorted[0][0],
                'fecha_fin'    : days_sorted[-1][0],
                'nivel'        : '🔴 CRÍTICO' if total_acct >= 2000
                                 else '🟠 ALTO'  if total_acct >= 1000
                                 else '🟡 MEDIO',
                'descripcion'  : f'Acumuló ${total_acct:.2f} en el período. '
                                 f'Cruzó ${SPIKE_THRESHOLD} el {dia_cruce} '
                                 f'(acumulado ese día: ${acum_cruce:.2f})'
            })

        # ── CASO 2: incremento entre 2 puntos cualquiera del período > $500 ──
        # Compara el acumulado en cada punto vs todos los puntos anteriores
        # Si la diferencia entre cualquier par (día_A, día_B) supera $500 → alerta
        alertas_cuenta = []
        for i in range(len(acumulados)):
            for j in range(i + 1, len(acumulados)):
                dia_a, _, acum_a = acumulados[i]
                dia_b, _, acum_b = acumulados[j]
                diff = acum_b - acum_a
                if diff >= SPIKE_THRESHOLD:
                    # Solo guardar el par con mayor incremento para esta cuenta
                    alertas_cuenta.append({
                        'account_id'   : acct_id,
                        'account_name' : name,
                        'dia_inicio'   : dia_a,
                        'dia_fin'      : dia_b,
                        'acum_inicio'  : acum_a,
                        'acum_fin'     : acum_b,
                        'incremento'   : round(diff, 2),
                        'nivel'        : '🔴 CRÍTICO' if diff >= 1000
                                         else '🟠 ALTO'  if diff >= 500
                                         else '🟡 MEDIO',
                        'descripcion'  : f'Del {dia_a} (acum ${acum_a:.2f}) al '
                                         f'{dia_b} (acum ${acum_b:.2f}): '
                                         f'+${diff:.2f}'
                    })

        # Guardar solo el par con el mayor incremento por cuenta
        if alertas_cuenta:
            mejor = max(alertas_cuenta, key=lambda x: x['incremento'])
            alerts_caso2.append(mejor)

    total_alertas = len(alerts_caso1) + len(alerts_caso2)

    return {
        'services'        : services,
        'by_account'      : dict(by_account),
        'trend'           : {k: dict(v) for k, v in trend.items()},
        'account_names'   : account_names,
        'daily_by_account': dict(daily_by_account),
        'alerts_caso1'    : alerts_caso1,    # Acumulado período cruza $500
        'alerts_caso2'    : alerts_caso2,    # Incremento entre 2 puntos > $500
        'total_alertas'   : total_alertas,
        'spike_threshold' : SPIKE_THRESHOLD,
        'total'           : round(total, 2),
        'period_start'    : start_date,
        'period_end'      : end_date,
    }

# ─────────────────────────────────────────────────────────────────
# 📊 EXCEL — 4 hojas
# ─────────────────────────────────────────────────────────────────
def build_excel(data, query_context=""):
    wb           = Workbook()
    services     = data['services']
    by_account   = data['by_account']
    trend        = data['trend']
    acc_names    = data['account_names']
    total        = data['total']
    period_start = data['period_start']
    period_end   = data['period_end']
    now_str      = datetime.utcnow().strftime('%d/%m/%Y %H:%M') + ' UTC'

    # ════════════════════════════════════════════════
    # 📋 HOJA 1 — Resumen Ejecutivo
    # ════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Resumen Ejecutivo"
    ws1.sheet_view.showGridLines = False

    ws1.merge_cells('A1:G2')
    ws1['A1'] = 'AWS Cost Report — Resumen Ejecutivo'
    ws1['A1'].font      = Font(name='Arial', bold=True, size=16, color=WHITE)
    ws1['A1'].fill      = fill(AWS_DARK)
    ws1['A1'].alignment = ctr()
    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 24

    ws1.merge_cells('A3:G3')
    ws1['A3'] = f'Generado: {now_str}  |  Período: {period_start} — {period_end}  |  Organización: Protecso'
    ws1['A3'].font      = Font(name='Arial', size=9, color=WHITE)
    ws1['A3'].fill      = fill(AWS_ORANGE)
    ws1['A3'].alignment = ctr()
    ws1.row_dimensions[3].height = 16

    tbl_start = 5
    if query_context:
        ws1.merge_cells('A4:G4')
        ws1['A4'] = f'Consulta: {query_context}'
        ws1['A4'].font      = Font(name='Arial', size=9, color=AMBER_C)
        ws1['A4'].fill      = fill(LIGHT_ORG)
        ws1['A4'].alignment = lft()
        ws1.row_dimensions[4].height = 16
        tbl_start = 6

    # 🏷 Cabecera tabla
    set_header_row(ws1, tbl_start,
        ['#', 'Servicio AWS', 'Costo (USD)', '% del Total', 'vs Promedio', 'Estado'],
        AWS_ORANGE)

    avg = total / len(services) if services else 0
    for idx, svc in enumerate(services):
        r   = tbl_start + 1 + idx
        bg  = WHITE if idx % 2 == 0 else LIGHT_BLUE
        pct = svc['cost'] / total if total else 0
        vs  = svc['cost'] / avg   if avg   else 0
        est = '🔴 Revisar'    if pct > 0.25 else \
              '⚠️ Monitorear' if pct > 0.10 else '✅ Normal'
        ws1.row_dimensions[r].height = 18
        for col, val in enumerate([idx+1, svc['service'], svc['cost'], pct, vs, est], 1):
            c = ws1.cell(r, col, val)
            c.fill  = fill(bg)
            c.border = bdr()
            c.font  = nf()
            if col == 3:
                c.number_format = '$#,##0.00'; c.alignment = rgt()
            elif col == 4:
                c.number_format = '0.0%';      c.alignment = rgt()
            elif col == 5:
                c.number_format = '0.00x';     c.alignment = rgt()
                if vs > 1.5: c.font = nf(bold=True, color=RED_C)
            elif col == 6:
                c.alignment = ctr()
                if '🔴' in str(val): c.font = nf(bold=True, color=RED_C)
                elif '⚠️' in str(val): c.font = nf(bold=True, color=AMBER_C)
            else:
                c.alignment = ctr() if col == 1 else lft()

    # 💰 Fila total
    r_tot = tbl_start + 1 + len(services)
    ws1.row_dimensions[r_tot].height = 22
    ws1.merge_cells(f'A{r_tot}:B{r_tot}')
    ws1[f'A{r_tot}'] = 'TOTAL PERÍODO'
    ws1[f'A{r_tot}'].font = hf(10); ws1[f'A{r_tot}'].fill = fill(AWS_DARK)
    ws1[f'A{r_tot}'].alignment = ctr(); ws1[f'A{r_tot}'].border = bdr()
    ws1[f'C{r_tot}'] = total
    ws1[f'C{r_tot}'].number_format = '$#,##0.00'
    ws1[f'C{r_tot}'].font = Font(name='Arial', bold=True, size=11, color=WHITE)
    ws1[f'C{r_tot}'].fill = fill(AWS_DARK); ws1[f'C{r_tot}'].alignment = rgt()
    ws1[f'C{r_tot}'].border = bdr()
    for col in [4, 5, 6]:
        ws1.cell(r_tot, col).fill = fill(AWS_DARK); ws1.cell(r_tot, col).border = bdr()

    ws1.column_dimensions['A'].width = 4
    ws1.column_dimensions['B'].width = 40
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 16
    ws1.column_dimensions['G'].width = 14

    # ════════════════════════════════════════════════
    # 🏢 HOJA 2 — Detalle por Cuenta
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detalle por Cuenta")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells('A1:E2')
    ws2['A1'] = 'Detalle de Costos por Cuenta AWS'
    ws2['A1'].font = Font(name='Arial', bold=True, size=14, color=WHITE)
    ws2['A1'].fill = fill(AWS_DARK); ws2['A1'].alignment = ctr()
    ws2.row_dimensions[1].height = 22; ws2.row_dimensions[2].height = 22

    row = 4
    for acct_id, svcs_list in sorted(by_account.items(),
                                      key=lambda x: sum(s['cost'] for s in x[1]),
                                      reverse=True):
        name       = acc_names.get(acct_id, acct_id)
        acct_total = sum(s['cost'] for s in svcs_list)

        ws2.merge_cells(f'A{row}:E{row}')
        ws2[f'A{row}'] = f'Cuenta: {name}  ({acct_id})  —  Total: ${acct_total:,.2f} USD'
        ws2[f'A{row}'].font = hf(10); ws2[f'A{row}'].fill = fill(HEADER2)
        ws2[f'A{row}'].alignment = lft(); ws2[f'A{row}'].border = bdr()
        ws2.row_dimensions[row].height = 20
        row += 1

        set_header_row(ws2, row, ['#', 'Servicio', 'Costo (USD)', '% Cuenta', '% Global'], AWS_ORANGE)
        row += 1

        for i, svc in enumerate(sorted(svcs_list, key=lambda x: x['cost'], reverse=True)):
            bg         = WHITE if i % 2 == 0 else LIGHT_GRAY
            pct_acct   = svc['cost'] / acct_total if acct_total else 0
            pct_global = svc['cost'] / total      if total      else 0
            set_data_row(ws2, row, [i+1, svc['service'], svc['cost'], pct_acct, pct_global],
                         bg, fmt_map={3: '$#,##0.00', 4: '0.0%', 5: '0.0%'})
            row += 1

        # 🔢 Subtotal por cuenta
        ws2.row_dimensions[row].height = 18
        ws2.merge_cells(f'A{row}:B{row}')
        ws2[f'A{row}'] = 'Subtotal'
        ws2[f'A{row}'].font = hf(9); ws2[f'A{row}'].fill = fill(AWS_ORANGE)
        ws2[f'A{row}'].alignment = rgt(); ws2[f'A{row}'].border = bdr()
        ws2[f'C{row}'] = acct_total
        ws2[f'C{row}'].number_format = '$#,##0.00'
        ws2[f'C{row}'].font = Font(name='Arial', bold=True, size=10, color=WHITE)
        ws2[f'C{row}'].fill = fill(AWS_ORANGE); ws2[f'C{row}'].alignment = rgt()
        ws2[f'C{row}'].border = bdr()
        for col in [4, 5]:
            ws2.cell(row, col).fill = fill(AWS_ORANGE); ws2.cell(row, col).border = bdr()
        row += 2

    ws2.column_dimensions['A'].width = 4
    ws2.column_dimensions['B'].width = 40
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 12

    # ════════════════════════════════════════════════
    # 🔍 HOJA 3 — Alertas de Consumo +$500 (2 casos)
    # ════════════════════════════════════════════════
    alerts_caso1  = data.get('alerts_caso1', [])
    alerts_caso2  = data.get('alerts_caso2', [])
    total_alertas = data.get('total_alertas', 0)
    threshold     = data.get('spike_threshold', 500)

    ws3 = wb.create_sheet("Alertas Consumo +$500")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells('A1:F2')
    ws3['A1'] = f'Alertas de Consumo Anómalo — Umbral +${threshold:,.0f} USD'
    ws3['A1'].font = Font(name='Arial', bold=True, size=14, color=WHITE)
    ws3['A1'].fill = fill(AWS_DARK); ws3['A1'].alignment = ctr()
    ws3.row_dimensions[1].height = 24; ws3.row_dimensions[2].height = 24

    ws3.merge_cells('A3:F3')
    ws3['A3'] = f'Período: {period_start} — {period_end}   |   Total alertas: {total_alertas}'
    ws3['A3'].font = Font(name='Arial', size=9, color=WHITE)
    ws3['A3'].fill = fill(AWS_ORANGE); ws3['A3'].alignment = ctr()
    ws3.row_dimensions[3].height = 16

    ws3.row_dimensions[5].height = 12
    ws3.row_dimensions[6].height = 30
    kpis = [
        ('CASO 1\nACUMULADO +$500', str(len(alerts_caso1)), RED_C),
        ('CASO 2\nINCREMENTO +$500', str(len(alerts_caso2)), AMBER_C),
        ('TOTAL\nALERTAS',           str(total_alertas),     "C0392B"),
        ('UMBRAL\nDETECCIÓN',       f'${threshold:,.0f}',   AWS_DARK),
        ('GASTO TOTAL\nPERÍODO',    f'${total:,.2f}',       AWS_DARK),
    ]
    for col, (label, value, color) in enumerate(kpis, 1):
        lc = ws3.cell(5, col, label)
        lc.font = Font(name='Arial', size=7, bold=True, color='888888')
        lc.alignment = ctr()
        vc = ws3.cell(6, col, value)
        vc.font = Font(name='Arial', size=14, bold=True, color=color)
        vc.fill = fill('F8F9FA'); vc.alignment = ctr(); vc.border = bdr()

    row = 8

    # 🔴 CASO 1 — Acumulado del período cruza $500
    ws3.merge_cells(f'A{row}:F{row}')
    ws3[f'A{row}'] = f'🔴 CASO 1 — Acumulado del Período supera ${threshold:,.0f} USD (día exacto del cruce)'
    ws3[f'A{row}'].font = Font(name='Arial', bold=True, size=11, color=WHITE)
    ws3[f'A{row}'].fill = fill('C0392B')
    ws3[f'A{row}'].alignment = lft(); ws3[f'A{row}'].border = bdr()
    ws3.row_dimensions[row].height = 20; row += 1

    if alerts_caso1:
        set_header_row(ws3, row,
            ['Cuenta', 'ID Cuenta', 'Total Acumulado', 'Día que cruzó $500', 'Período', 'Nivel'],
            "C0392B", fsize=8)
        row += 1
        for a in sorted(alerts_caso1, key=lambda x: x['total'], reverse=True):
            bg = LIGHT_RED if a['total'] >= 2000 else LIGHT_ORG
            ws3.row_dimensions[row].height = 18
            vals = [a['account_name'], a['account_id'], a['total'],
                    a['dia_cruce'], f"{a['fecha_inicio']} → {a['fecha_fin']}", a['nivel']]
            for col, val in enumerate(vals, 1):
                c = ws3.cell(row, col, val)
                c.fill = fill(bg); c.border = bdr(); c.font = nf()
                if col == 3:
                    c.number_format = '$#,##0.00'; c.alignment = rgt()
                    c.font = nf(bold=True, color=RED_C if a['total'] >= 1000 else AMBER_C)
                else: c.alignment = lft()
            row += 1
    else:
        ws3.merge_cells(f'A{row}:F{row}')
        ws3[f'A{row}'] = f'✅ Ninguna cuenta superó ${threshold:,.0f} acumulado en el período'
        ws3[f'A{row}'].font = nf(color=GREEN_C); ws3[f'A{row}'].fill = fill(LIGHT_GRN)
        ws3[f'A{row}'].alignment = ctr(); ws3[f'A{row}'].border = bdr()
        row += 1

    row += 1

    # 🟠 CASO 2 — Incremento entre 2 puntos del período > $500
    ws3.merge_cells(f'A{row}:F{row}')
    ws3[f'A{row}'] = f'🟠 CASO 2 — Incremento entre 2 puntos del período supera ${threshold:,.0f} USD'
    ws3[f'A{row}'].font = Font(name='Arial', bold=True, size=11, color=WHITE)
    ws3[f'A{row}'].fill = fill(AMBER_C)
    ws3[f'A{row}'].alignment = lft(); ws3[f'A{row}'].border = bdr()
    ws3.row_dimensions[row].height = 20; row += 1

    if alerts_caso2:
        set_header_row(ws3, row,
            ['Cuenta', 'ID Cuenta', 'Desde (acum $)', 'Hasta (acum $)', 'Incremento +$', 'Nivel'],
            AMBER_C, fsize=8)
        row += 1
        for a in sorted(alerts_caso2, key=lambda x: x['incremento'], reverse=True):
            bg = LIGHT_RED if a['incremento'] >= 1000 else LIGHT_ORG
            ws3.row_dimensions[row].height = 18
            vals = [
                a['account_name'], a['account_id'],
                f"{a['dia_inicio']} (${a['acum_inicio']:,.2f})",
                f"{a['dia_fin']} (${a['acum_fin']:,.2f})",
                a['incremento'], a['nivel']
            ]
            for col, val in enumerate(vals, 1):
                c = ws3.cell(row, col, val)
                c.fill = fill(bg); c.border = bdr(); c.font = nf()
                if col == 5:
                    c.number_format = '$#,##0.00'; c.alignment = rgt()
                    c.font = nf(bold=True, color=RED_C if a['incremento'] >= 1000 else AMBER_C)
                else: c.alignment = lft()
            row += 1
    else:
        ws3.merge_cells(f'A{row}:F{row}')
        ws3[f'A{row}'] = f'✅ Sin incrementos mayores a ${threshold:,.0f} entre puntos del período'
        ws3[f'A{row}'].font = nf(color=GREEN_C); ws3[f'A{row}'].fill = fill(LIGHT_GRN)
        ws3[f'A{row}'].alignment = ctr(); ws3[f'A{row}'].border = bdr()
        row += 1

    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 24
    ws3.column_dimensions['D'].width = 24
    ws3.column_dimensions['E'].width = 16
    ws3.column_dimensions['F'].width = 16

        # 💾 Serializar a bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────────────────────────
# 📧 AWS SES — envío con adjunto Excel
# ─────────────────────────────────────────────────────────────────
def send_ses(to_email, excel_bytes, data, query_context="", note=""):
    period = f"{data['period_start']} — {data['period_end']}"
    total  = data['total']

    subject = f'AWS Cost Report | {period} | ${total:,.2f} USD'

    # 🧹 Limpiar Markdown del contexto para que no aparezcan ** ni ## en el email
    import re
    def clean_md(text):
        text = re.sub(r'\*\*(.+?)\*\*', r'\1', text)   # **negrita** → texto
        text = re.sub(r'\*(.+?)\*',     r'\1', text)   # *itálica* → texto
        text = re.sub(r'#{1,6}\s*',     '',    text)   # ## títulos → sin #
        text = re.sub(r'\n{3,}',        '\n\n', text)  # múltiples saltos → doble
        return text.strip()

    if query_context:
        query_context = clean_md(query_context)

    # 💬 Bloque de mensaje personalizado (NUEVO)
    note_block = ""
    if note:
        note_block = f"""
  <div style="background:#f0f4ff;border-left:4px solid #7c3aed;border-radius:0 8px 8px 0;
              padding:14px 18px;margin-bottom:20px;">
    <div style="font-size:11px;font-weight:bold;color:#7c3aed;text-transform:uppercase;
                letter-spacing:.08em;margin-bottom:6px;">Mensaje del remitente</div>
    <div style="font-size:14px;color:#232F3E;line-height:1.6;">{note}</div>
  </div>"""

    # 🔶 Bloque de contexto de consulta (si existe)
    context_block = ""
    if query_context:
        context_block = f"""
  <div style="background:#FEF9E7;border-left:4px solid #FF9900;
              padding:10px 14px;margin-bottom:16px;font-size:13px;color:#7D6608;">
    <strong>Contexto:</strong> {query_context}
  </div>"""

    html_body = f"""
<html><body style="font-family:Arial,sans-serif;color:#232F3E;max-width:600px;margin:0 auto">

  <!-- 🔝 Header -->
  <div style="background:#232F3E;padding:24px;text-align:center;border-radius:8px 8px 0 0">
    <h2 style="color:#FF9900;margin:0;font-size:22px">&#9729; AWS Cost Report</h2>
    <p style="color:#8ca0b8;margin:8px 0 0;font-size:13px">{period}</p>
  </div>

  <div style="background:#f8f9fa;padding:24px;border:1px solid #e0e0e0">

    {note_block}
    {context_block}

    <!-- 📊 KPIs -->
    <table style="width:100%;border-collapse:collapse;margin-bottom:20px">
      <tr>
        <td style="background:#fff;border:1px solid #e0e0e0;border-radius:8px;
                   padding:16px;text-align:center;width:33%">
          <div style="font-size:11px;color:#888">GASTO TOTAL</div>
          <div style="font-size:22px;font-weight:bold;color:#D68910">${total:,.2f}</div>
          <div style="font-size:11px;color:#888">USD</div>
        </td>
        <td style="width:2%"></td>
        <td style="background:#fff;border:1px solid #e0e0e0;border-radius:8px;
                   padding:16px;text-align:center;width:33%">
          <div style="font-size:11px;color:#888">SERVICIOS</div>
          <div style="font-size:22px;font-weight:bold;color:#1A5276">{len(data['services'])}</div>
          <div style="font-size:11px;color:#888">activos</div>
        </td>
        <td style="width:2%"></td>
        <td style="background:#fff;border:1px solid #e0e0e0;border-radius:8px;
                   padding:16px;text-align:center;width:30%">
          <div style="font-size:11px;color:#888">CUENTAS</div>
          <div style="font-size:22px;font-weight:bold;color:#117A65">{len(data['by_account'])}</div>
          <div style="font-size:11px;color:#888">con gasto</div>
        </td>
      </tr>
    </table>

    <!-- 🔥 Top 5 servicios -->
    <h3 style="color:#232F3E;font-size:14px;margin:0 0 10px">Top 5 servicios</h3>
    <table style="width:100%;border-collapse:collapse;font-size:13px">
      <tr style="background:#232F3E;color:#FF9900">
        <th style="padding:8px 10px;text-align:left">#</th>
        <th style="padding:8px 10px;text-align:left">Servicio</th>
        <th style="padding:8px 10px;text-align:right">Costo (USD)</th>
        <th style="padding:8px 10px;text-align:right">% Total</th>
      </tr>
      {''.join(
        f'<tr style="background:{"#f8f9fa" if i%2==0 else "#fff"}">'
        f'<td style="padding:7px 10px;color:#888">{i+1}</td>'
        f'<td style="padding:7px 10px">{s["service"]}</td>'
        f'<td style="padding:7px 10px;text-align:right;font-weight:bold">${s["cost"]:,.2f}</td>'
        f'<td style="padding:7px 10px;text-align:right;color:{"#C0392B" if s["cost"]/total>0.25 else "#555"}">'
        f'{s["cost"]/total*100:.1f}%</td></tr>'
        for i, s in enumerate(data['services'][:5])
      )}
    </table>

    <p style="font-size:12px;color:#888;margin-top:20px">
      Excel adjunto con 3 hojas: <strong>Resumen Ejecutivo</strong> ·
      <strong>Detalle por Cuenta</strong> · <strong>Alertas Consumo +$500</strong>
    </p>
  </div>

  <!-- 🔻 Footer -->
  <div style="background:#232F3E;padding:12px;text-align:center;border-radius:0 0 8px 8px">
    <p style="color:#8ca0b8;font-size:11px;margin:0">
      Generado automáticamente por AWS Cost Assistant · Protecso
    </p>
  </div>

</body></html>
"""

    # 📨 Armar mensaje MIME con adjunto
    msg = mp.MIMEMultipart('mixed')
    msg['Subject'] = subject
    msg['From']    = f'AWS Cost Assistant <{FROM_EMAIL}>'
    msg['To']      = to_email

    alt = mp.MIMEMultipart('alternative')
    alt.attach(mt.MIMEText(html_body, 'html', 'utf-8'))
    msg.attach(alt)

    # 📎 Adjuntar Excel
    fname = f'AWS_Cost_Report_{data["period_start"]}_{data["period_end"]}.xlsx'
    att   = ma.MIMEApplication(
        excel_bytes,
        _subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    att.add_header('Content-Disposition', 'attachment', filename=fname)
    msg.attach(att)

    ses_client.send_raw_email(
        Source=FROM_EMAIL,
        Destinations=[to_email],
        RawMessage={'Data': msg.as_string()}
    )
    return fname

# ─────────────────────────────────────────────────────────────────
# 🚀 HANDLER PRINCIPAL
# ─────────────────────────────────────────────────────────────────
def lambda_handler(event, context):
    print("Event:", json.dumps(event))

    # 🌐 CORS preflight
    if event.get('httpMethod') == 'OPTIONS':
        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin' : '*',
                'Access-Control-Allow-Headers': 'Content-Type,Authorization',
                'Access-Control-Allow-Methods': 'POST,OPTIONS'
            },
            'body': ''
        }

    try:
        body          = json.loads(event.get('body', '{}'))
        to_email      = body.get('email', '').strip()
        days          = int(body.get('days', 30))
        query_context = body.get('query_context', '')
        note          = body.get('note', '').strip()   # ← 💬 mensaje personalizado
        start_date    = body.get('start_date', None)   # ← 📅 fecha inicio exacta (NUEVO)
        end_date      = body.get('end_date',   None)   # ← 📅 fecha fin exacta   (NUEVO)

        # ❌ Validación básica
        if not to_email or '@' not in to_email:
            return {
                'statusCode': 400,
                'headers': {'Access-Control-Allow-Origin': '*'},
                'body': json.dumps({'error': 'Correo inválido o faltante'})
            }

        # 1️⃣ Consultar datos desde Athena
        data = fetch_costs(days, start_date, end_date)

        # 2️⃣ Generar Excel con 4 hojas
        excel_bytes = build_excel(data, query_context)

        # 3️⃣ Enviar por SES con mensaje personalizado
        fname = send_ses(to_email, excel_bytes, data, query_context, note)

        return {
            'statusCode': 200,
            'headers': {
                'Access-Control-Allow-Origin': '*',
                'Content-Type': 'application/json'
            },
            'body': json.dumps({
                'success' : True,
                'message' : f'Reporte enviado a {to_email}',
                'filename': fname,
                'total'   : data['total'],
                'period'  : f"{data['period_start']} — {data['period_end']}",
                'services': len(data['services']),
                'accounts': len(data['by_account'])
            }, ensure_ascii=False)
        }

    except Exception as e:
        print(f'❌ ERROR: {e}')
        import traceback; traceback.print_exc()
        return {
            'statusCode': 500,
            'headers': {'Access-Control-Allow-Origin': '*'},
            'body': json.dumps({'error': str(e)})
        }
